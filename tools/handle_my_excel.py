import os
from openpyxl import load_workbook


class MyExcel:
    def __init__(self, excel_path: str, sheet_name):
        """
        1、需要判断路径是否存在。如果不存在，抛异常。如果存在，打开。
        :param excel_path: 完整的excel文件路径
        """
        if os.path.isfile(excel_path):
            if not excel_path.endswith(".xlsx"):
                print("文件不是以xlsx结尾，不支持处理。")
            else:
                self.wb = load_workbook(excel_path)
                self.select_sheet_by_name(sheet_name)
        else:
            print("文件路径不存在。")
            raise FileNotFoundError

    def select_sheet_by_name(self, name):
        if name not in self.wb.sheetnames:
            print("表单名称不存在！")
            self.sh = None
        else:
            self.sh = self.wb[name]

    def read_all_datas(self):
        if self.sh:
            cases_list = []
            # 获取表单里的所有数据
            all_values = list(self.sh.values)
            # 获取表单第一行，作为key
            keys = all_values[0]
            # 遍历表单，从第2行开始
            for values in all_values[1:]:
                # 把key和每一行进行组合成字典
                case = dict(zip(keys, values))
                cases_list.append(case)
            return cases_list
